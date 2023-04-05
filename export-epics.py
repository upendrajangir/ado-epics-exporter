import requests
import base64
import json
import os
import logging
import coloredlogs
import openpyxl
import smtplib
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Set logging
logger = logging.getLogger(__name__)
coloredlogs.install(
    fmt="%(asctime)s | %(hostname)s | %(levelname)s | %(name)s | %(filename)s:%(lineno)d | %(message)s",
    level="WARNING",
)


def get_epics(organization: str, project: str, personal_access_token: str) -> List[str]:
    """
    Get a list of IDs for all epics in the specified Azure DevOps organization and project.

    Args:
        organization (str): The name of the Azure DevOps organization.
        project (str): The name of the project within the organization.
        personal_access_token (str): The personal access token for authenticating to the Azure DevOps API.

    Returns:
        List[str]: A list of IDs for all epics in the specified project.

    Raises:
        Exception: If the API request fails for any reason.
    """
    if not organization or not project or not personal_access_token:
        raise ValueError("Required parameters not provided")

    url = (
        f"https://dev.azure.com/{organization}/{project}/_apis/wit/wiql?api-version=6.0"
    )
    encoded_pat = base64.b64encode((":" + personal_access_token).encode()).decode()
    headers = {"Authorization": "Basic " + encoded_pat}
    query = {
        "query": "SELECT [System.Id] FROM workitems WHERE [System.WorkItemType] = 'Epic'"
    }

    try:
        response = requests.post(url, headers=headers, json=query, timeout=30)
        response.raise_for_status()
    except requests.exceptions.JSONDecodeError as JSON_err:
        logging.error(f"Error decoding JSON response: {JSON_err}")
        raise requests.exceptions.JSONDecodeError(
            f"Error decoding JSON response: {JSON_err}"
        )
        return []
    except requests.exceptions.HTTPError as http_err:
        logging.error(f"HTTP Error: {http_err}\n{response.text}")
        raise requests.exceptions.HTTPError(f"HTTP Error: {http_err}\n{response.text}")
    except requests.exceptions.ConnectionError as conn_err:
        logging.error(f"Error Connecting: {conn_err}")
        raise requests.exceptions.ConnectionError(f"Error Connecting: {conn_err}")
    except requests.exceptions.Timeout as timeout_err:
        logging.error(f"Timeout Error: {timeout_err}")
        raise requests.exceptions.Timeout(f"Timeout Error: {timeout_err}")
    except requests.exceptions.RequestException as req_err:
        logging.error(f"Request Exception: {req_err}")
        raise requests.exceptions.RequestException(f"Request Exception: {req_err}")

    data = response.json()
    work_items = [work_item["id"] for work_item in data["workItems"]]
    return work_items


def get_work_items(
    epic_ids: List[str],
    organization: str,
    project: str,
    personal_access_token: str,
    api_version: str,
) -> List[Dict]:
    """
    Get a list of work items for the given epic IDs in the specified Azure DevOps organization and project.

    Args:
        epic_ids (List[str]): A list of IDs for the epics whose work items are to be fetched.
        organization (str): The name of the Azure DevOps organization.
        project (str): The name of the project within the organization.
        personal_access_token (str): The personal access token for authenticating to the Azure DevOps API.
        api_version (str): The version of the Azure DevOps API to use.

    Returns:
        List[Dict]: A list of work item data for the specified epics.

    Raises:
        ValueError: If any required parameters are not provided.
        requests.exceptions.RequestException: If an error occurs while making the API request.
    """
    if not organization or not project or not personal_access_token or not epic_ids:
        raise ValueError("Required parameters not provided")

    work_items = []
    encoded_pat = base64.b64encode((":" + personal_access_token).encode()).decode()
    headers = {"Authorization": "Basic " + encoded_pat}

    for epic_id in epic_ids:
        url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/workitems/{epic_id}?api-version={api_version}&$expand=all"

        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
        except requests.exceptions.HTTPError as http_err:
            print(f"HTTP Error: {http_err}")
            print(response.text)
            continue
        except requests.exceptions.ConnectionError as conn_err:
            print(f"Error Connecting: {conn_err}")
            continue
        except requests.exceptions.Timeout as timeout_err:
            print(f"Timeout Error: {timeout_err}")
            continue
        except requests.exceptions.RequestException as req_err:
            print(f"Request Exception: {req_err}")
            continue

        data = response.json()
        work_items.append(data)

    return work_items


# Function to create an Excel file with the work items data
def create_excel(work_items):
    work_items_data = []

    for work_item_id in work_items:
        url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/workItems/{work_item_id}?api-version={api_version}"
        headers = {"Authorization": "Basic " + personal_access_token}
        response = requests.get(url, headers=headers)
        data = response.json()

        work_items_data.append(
            {
                "ID": data["id"],
                "Title": data["fields"]["System.Title"],
                "Assigned To": data["fields"]["System.AssignedTo"]["displayName"],
                "State": data["fields"]["System.State"],
                "Area Path": data["fields"]["System.AreaPath"],
                "Iteration Path": data["fields"]["System.IterationPath"],
            }
        )

    df = pd.DataFrame(work_items_data)
    excel_file = "Epic_Work_Items.xlsx"
    df.to_excel(excel_file, index=False)

    return excel_file


def write_epics_to_excel(epic_list: List[Dict]) -> None:
    workbook = openpyxl.Workbook()
    del workbook["Sheet"]

    # Create sheets for each work item state
    todo_sheet = workbook.create_sheet("To Do")
    doing_sheet = workbook.create_sheet("Doing")
    done_sheet = workbook.create_sheet("Done")

    state_sheets = {
        "To Do": todo_sheet,
        "Doing": doing_sheet,
        "Done": done_sheet,
    }

    # Formatting
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    header_alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # Function to write header row to a sheet
    def write_header(sheet):
        headers = [
            "ID",
            "Title",
            "State",
            "Priority",
            "Start Date",
            "Target Date",
            "Description",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Write header rows to each sheet
    for sheet in state_sheets.values():
        write_header(sheet)

    # Convert the RevisedDate string to a datetime object if it's a string
    for epic in epic_list:
        if isinstance(epic["fields"]["System.RevisedDate"], str):
            epic["fields"]["System.RevisedDate"] = datetime.fromisoformat(
                epic["fields"]["System.RevisedDate"].rstrip("Z")
            )

    # Sort the epics by RevisedDate in descending order
    sorted_epics = sorted(
        epic_list, key=lambda x: x["fields"]["System.RevisedDate"], reverse=True
    )

    # Write data rows to respective sheets based on work item state
    row_counters = {"To Do": 2, "Doing": 2, "Done": 2}

    for epic in sorted_epics:
        state = epic["fields"]["System.State"]
        sheet = state_sheets[state]
        row = row_counters[state]

        sheet.cell(row=row, column=1, value=epic.get("id"))
        sheet.cell(row=row, column=2, value=epic.get("fields", {}).get("System.Title"))
        sheet.cell(row=row, column=3, value=epic.get("fields", {}).get("System.State"))
        sheet.cell(
            row=row,
            column=4,
            value=epic.get("fields", {}).get("Microsoft.VSTS.Common.Priority", ""),
        )
        sheet.cell(
            row=row,
            column=5,
            value=epic.get("fields", {}).get("Microsoft.VSTS.Scheduling.StartDate", ""),
        )
        sheet.cell(
            row=row,
            column=6,
            value=epic.get("fields", {}).get(
                "Microsoft.VSTS.Scheduling.TargetDate", ""
            ),
        )
        sheet.cell(
            row=row,
            column=7,
            value=epic.get("fields", {}).get("System.Description", ""),
        )

        row_counters[state] += 1

    # Adjust column widths for each sheet
    for sheet in state_sheets.values():
        for col in range(1, 8):
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].auto_size = True

    # Save the Excel file
    workbook.save("epics.xlsx")
    return ("epics.xlsx")


# Function to send an email with the Excel attachment
def send_email(
    excel_file: str,
    sender_email: str,
    receiver_emails: List[str],
    email_password: str,
    smtp_server: str,
    smtp_port: int,
) -> str:
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = ", ".join(receiver_emails)
    today_date = datetime.now().strftime("%Y-%m-%d")
    msg["Subject"] = f"Epic Work Items Report for {today_date}"

    body = f"""
    <html>
        <body>
            <h2>Epic Work Items Report for {today_date}</h2>
            <p>Dear Team,</p>
            <p>Please find the attached <strong>Epic Work Items Report</strong> for {today_date}. This report provides a detailed overview of the progress of various epic work items, including their states and priorities.</p>
            <p>We encourage you to review the report and use the insights to plan your tasks and priorities effectively. If you have any questions or need further clarification, please feel free to reach out to the project manager.</p>
            <p>Best regards,</p>
            <p>Your Team</p>
        </body>
    </html>
    """

    msg.attach(MIMEText(body, "html"))

    with open(excel_file, "rb") as attachment:
        base = MIMEBase("application", "octet-stream")
        base.set_payload(attachment.read())
        encoders.encode_base64(base)
        base.add_header("Content-Disposition", f"attachment; filename={excel_file}")
        msg.attach(base)

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, email_password)
        text = msg.as_string()
        server.sendmail(sender_email, receiver_emails, text)
        server.quit()
        return f"Email sent successfully to: {', '.join(receiver_emails)}"
    except Exception as e:
        return f"Error sending email: {str(e)}"


# Main script
if __name__ == "__main__":
    # Azure DevOps API settings
    organization = "DevOpsCelebal"
    project = "Chitransh"
    api_version = "7.1-preview.3"
    personal_access_token = "yfe5men7x42qy7xk4zqf4gyzmg6hcpbrzg4tpowt75xvjplgsdga"

    # Email settings
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "you@gmail.com"
    receiver_emails = ["person1@example.com", "person2@example.com"]
    email_password = "your_email_password"

    # Get the list of epics
    epics = get_epics(
        organization=organization,
        project=project,
        personal_access_token=personal_access_token,
    )

    # Get the list of work items
    work_items = get_work_items(
        epic_ids=epics,
        organization=organization,
        project=project,
        personal_access_token=personal_access_token,
        api_version=api_version,
    )

    # Write the work items to an Excel file
    excel_file_path = write_epics_to_excel(epic_list=work_items)

    # Send the Excel file as an email attachment
    send_email(excel_file_path, sender_email, receiver_emails, email_password)

    # Delete the Excel file
    os.remove(excel_file_path)
